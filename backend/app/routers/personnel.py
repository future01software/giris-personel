from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models import PersonnelCreate, BulkDeleteRequest
from app.db import db, DEMO_MODE
from app.deps import get_current_user, require_role
from app.utils import new_id, compute_can_enter_map, prepare_turkish_search, parse_dt_safe

router = APIRouter(prefix="/personnel", tags=["personnel"])
TR_TZ = ZoneInfo("Europe/Istanbul")


def _excel_text(value) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _excel_date(value):
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None)
    except (TypeError, ValueError):
        return _excel_text(value)


def _log_timestamp(log: dict) -> float:
    for key in ("timestamp_ts", "created_at_ts"):
        value = log.get(key)
        if isinstance(value, (int, float)):
            return float(value)
    for key in ("timestamp", "created_at"):
        value = log.get(key)
        if value:
            try:
                return datetime.fromisoformat(str(value).replace("Z", "+00:00")).timestamp()
            except (TypeError, ValueError):
                pass
    return 0.0


def _log_action(log: dict) -> str:
    value = str(log.get("action") or log.get("decision") or "").strip().upper()
    if value in ("IN", "APPROVED", "ALLOW", "ALLOWED", "ACCEPTED", "OK"):
        return "IN"
    if value in ("OUT", "REJECTED", "DENY", "DENIED", "NOT_OK", "NO"):
        return "OUT"
    return ""


async def _activity_summary(personnel_id: str) -> dict:
    query = {"$or": [{"personnel_id": personnel_id}, {"person_id": personnel_id}]}
    logs = await db.entry_logs.find(query, {"_id": 0}).to_list(5000)
    logs.sort(key=_log_timestamp)

    now_ts = datetime.now(timezone.utc).timestamp()
    today_start = datetime.now(TR_TZ).replace(hour=0, minute=0, second=0, microsecond=0).timestamp()
    last_entry = None
    last_exit = None
    last_gate = None
    last_gate_ts = 0.0
    open_entry_ts = None
    today_total_sec = 0

    for log in logs:
        ts = _log_timestamp(log)
        action = _log_action(log)
        if not ts or not action:
            continue

        gate = log.get("gate") or log.get("security_unit")
        if gate and ts >= last_gate_ts:
            last_gate = gate
            last_gate_ts = ts

        if action == "IN":
            last_entry = ts
            open_entry_ts = ts
        elif action == "OUT":
            last_exit = ts
            if open_entry_ts is not None:
                overlap_start = max(open_entry_ts, today_start)
                overlap_end = min(ts, now_ts)
                if overlap_end > overlap_start:
                    today_total_sec += int(overlap_end - overlap_start)
            open_entry_ts = None

    if open_entry_ts is not None:
        overlap_start = max(open_entry_ts, today_start)
        if now_ts > overlap_start:
            today_total_sec += int(now_ts - overlap_start)

    def iso(ts):
        return datetime.fromtimestamp(ts, timezone.utc).isoformat() if ts else None

    return {
        "last_entry_at": iso(last_entry),
        "last_exit_at": iso(last_exit),
        "today_inside_seconds": today_total_sec,
        "last_gate": last_gate,
    }


@router.post("")
async def create_personnel(data: PersonnelCreate, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["admin"])

    personnel_doc = {
        "id": new_id("personnel"),
        "full_name": data.full_name,
        "tc_number": data.tc_number,
        "company": data.company,
        "phone": data.phone,
        "license_plate": data.license_plate,
        "photo_url": data.photo_url,
        "assignment_start": data.assignment_start,
        "assignment_end": data.assignment_end,
        "entry_blocked": data.entry_blocked,
        "entry_block_reason": data.entry_block_reason,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.personnel.insert_one(personnel_doc)
    return {"message": "Personnel created", "id": personnel_doc["id"]}


@router.get("")
async def get_personnel(
    page: int = 1,
    limit: int = 50,
    search: str = None,
    status: str = None,   # "can" | "cannot" | "all"
    company: str = None,  # ✅ EKLENDI
    current_user: dict = Depends(get_current_user),
):
    if DEMO_MODE:
        mock_data = [
            {
                "id": f"demo_p_{i}",
                "full_name": name,
                "tc_number": f"123456789{i:02d}",
                "company": "Demolife İnşaat",
                "phone": "+90 555 111 22 33",
                "can_enter": True if i % 4 != 0 else False,
                "assignment_end": (datetime.now() + timedelta(days=30)).isoformat()
            }
            for i, name in enumerate(["Ahmet Yılmaz", "Mehmet Demir", "Ayşe Kaya", "Fatma Çelik", "Can Özkan"])
        ]
        return {
            "data": mock_data,
            "total": len(mock_data),
            "page": 1,
            "limit": 50,
            "pages": 1
        }

    status = (status or "").strip().lower()
    if status not in ["", "all", "can", "cannot"]:
        status = ""

    query = {}

    # ✅ Company filtresi (backend pagination doğru çalışsın diye)
    if company:
        company = company.strip()
        if company:
            query["company"] = company

    if search:
        # Search varsa, özel regex fonksiyonunu kullan (fonksiyon yukarıda tanımlandı kabul ediyoruz, 
        # aslında aynı dosyada ama search_personnel içinde tanımlamıştık, dışarı almamız lazım.
        # Düzeltme: helper'ı yukarıda global tanımladım, burada sadece çağırıyorum.
        regex_pattern = prepare_turkish_search(search)
        query["$or"] = [
            {"full_name": {"$regex": regex_pattern, "$options": "i"}},
            {"tc_number": {"$regex": regex_pattern, "$options": "i"}},
            {"company": {"$regex": regex_pattern, "$options": "i"}},
        ]

    if status in ["", "all"]:
        skip = (page - 1) * limit
        total = await db.personnel.count_documents(query)
        
        # Projection to limit fields returned
        projection = {
            "_id": 0,
            "id": 1,
            "full_name": 1,
            "tc_number": 1,
            "company": 1,
            "phone": 1,
            "license_plate": 1,
            "photo_url": 1,
            "assignment_start": 1,
            "assignment_end": 1,
            "entry_blocked": 1,
            "entry_block_reason": 1,
            "created_at": 1
        }
        
        personnel_list = await db.personnel.find(query, projection).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

        can_map = await compute_can_enter_map(personnel_list)
        for p in personnel_list:
            pid = p.get("id")
            p["can_enter"] = bool(can_map.get(pid, True))

        return {
            "data": personnel_list,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit
        }

    # status=can/cannot
    all_candidates = await db.personnel.find(
        query,
        {"_id": 0, "id": 1, "assignment_start": 1, "assignment_end": 1, "entry_blocked": 1}
    ).sort("created_at", -1).to_list(200000)
    can_map = await compute_can_enter_map(all_candidates)

    want = True if status == "can" else False
    filtered_ids = [p["id"] for p in all_candidates if can_map.get(p["id"], True) == want]

    total = len(filtered_ids)
    pages = (total + limit - 1) // limit if total > 0 else 1

    start = (page - 1) * limit
    end = start + limit
    page_ids = filtered_ids[start:end]

    if not page_ids:
        return {"data": [], "total": total, "page": page, "limit": limit, "pages": pages}

    personnel_list = await db.personnel.find({"id": {"$in": page_ids}}, {"_id": 0}).to_list(limit)

    order = {pid: i for i, pid in enumerate(page_ids)}
    personnel_list.sort(key=lambda x: order.get(x.get("id"), 10**9))

    for p in personnel_list:
        pid = p.get("id")
        p["can_enter"] = bool(can_map.get(pid, True))

    return {"data": personnel_list, "total": total, "page": page, "limit": limit, "pages": pages}


@router.get("/search")
async def search_personnel(
    q: str = None, 
    name: str = None, 
    surname: str = None, 
    tc: str = None, 
    current_user: dict = Depends(get_current_user)
):
    if DEMO_MODE:
        demo_personnel = [
            {"id": "demo_p_0", "full_name": "Ahmet Yılmaz", "tc_number": "12345678900", "company": "Demolife İnşaat", "phone": "+90 555 111 22 33", "overall_status": "green", "assignment_end": (datetime.now() + timedelta(days=60)).isoformat()},
            {"id": "demo_p_1", "full_name": "Mehmet Demir", "tc_number": "12345678901", "company": "Demolife İnşaat", "phone": "+90 555 222 33 44", "overall_status": "green", "assignment_end": (datetime.now() + timedelta(days=45)).isoformat()},
            {"id": "demo_p_2", "full_name": "Ayşe Kaya", "tc_number": "12345678902", "company": "Yıldız Liman", "phone": "+90 555 333 44 55", "overall_status": "yellow", "assignment_end": (datetime.now() + timedelta(days=10)).isoformat()},
            {"id": "demo_p_3", "full_name": "Fatma Çelik", "tc_number": "12345678903", "company": "Yıldız Liman", "phone": "+90 555 444 55 66", "overall_status": "red", "assignment_end": (datetime.now() - timedelta(days=5)).isoformat()},
            {"id": "demo_p_4", "full_name": "Can Özkan", "tc_number": "12345678904", "company": "Atlas Tersane", "phone": "+90 555 555 66 77", "overall_status": "green", "assignment_end": (datetime.now() + timedelta(days=90)).isoformat()},
        ]
        search_term = (q or name or surname or tc or "").lower()
        if search_term:
            return [p for p in demo_personnel if search_term in p["full_name"].lower() or search_term in p["tc_number"] or search_term in p["company"].lower()]
        return demo_personnel

    clauses = []

    # 1) Eski genel arama (q)
    if q:
        regex_q = prepare_turkish_search(q)
        clauses.append({
            "$or": [
                {"full_name": {"$regex": regex_q, "$options": "i"}},
                {"tc_number": {"$regex": regex_q, "$options": "i"}},
                {"company": {"$regex": regex_q, "$options": "i"}},
                {"license_plate": {"$regex": regex_q, "$options": "i"}},
            ]
        })

    # 2) Gelişmiş arama alanları (name, surname, tc)
    if name:
        regex_name = prepare_turkish_search(name)
        clauses.append({"full_name": {"$regex": regex_name, "$options": "i"}})
    
    if surname:
        regex_surname = prepare_turkish_search(surname)
        clauses.append({"full_name": {"$regex": regex_surname, "$options": "i"}})

    if tc:
        regex_tc = prepare_turkish_search(tc)
        clauses.append({"tc_number": {"$regex": regex_tc, "$options": "i"}})

    if not clauses:
        return []

    query = {"$and": clauses} if len(clauses) > 1 else clauses[0]
    
    results = await db.personnel.find(query, {"_id": 0}).to_list(100)
    return results


@router.get("/companies")
async def get_personnel_companies(current_user: dict = Depends(get_current_user)):
    companies = await db.personnel.distinct("company")
    companies = [str(c).strip() for c in companies if c and str(c).strip()]
    companies.sort()
    return {"companies": companies}


@router.get("/export/excel")
async def export_personnel_excel(current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["admin"])

    personnel = await db.personnel.find({}, {"_id": 0}).sort("full_name", 1).to_list(200000)
    document_types = await db.document_types.find({}, {"_id": 0}).sort("name_tr", 1).to_list(1000)
    documents = await db.personnel_documents.find({}, {"_id": 0}).to_list(500000)

    document_by_person = {}
    for document in documents:
        personnel_id = document.get("personnel_id")
        document_type_id = document.get("document_type_id")
        if not personnel_id or not document_type_id:
            continue

        key = (personnel_id, document_type_id)
        current = document_by_person.get(key)
        if current is None or str(document.get("expiry_date") or "") > str(current.get("expiry_date") or ""):
            document_by_person[key] = document

    can_enter_map = await compute_can_enter_map(personnel)
    mandatory_document_types = [item for item in document_types if item.get("is_mandatory")]
    now = datetime.now(timezone.utc)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Personel ve Evraklar"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"

    export_time = datetime.now(TR_TZ)
    sheet["A1"] = "Clear2Work Personel ve Evrak Takip Yedeği"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B4778")
    sheet["A2"] = f"Oluşturulma: {export_time.strftime('%d.%m.%Y %H:%M')} — Giriş/çıkış kayıtları dahil değildir."
    sheet["A2"].font = Font(size=10, italic=True, color="52677D")

    base_headers = [
        "Ad Soyad",
        "TC Kimlik No",
        "Firma",
        "Telefon",
        "Plaka",
        "Görevlendirme Başlangıç",
        "Görevlendirme Bitiş",
        "Giriş Yetkisi",
        "Engel Nedeni",
    ]
    headers = list(base_headers)
    for document_type in document_types:
        name = document_type.get("name_tr") or document_type.get("name_en") or "Evrak"
        headers.extend([f"{name} Son Geçerlilik", f"{name} Kalan Gün", f"{name} Durum"])

    last_column = get_column_letter(len(headers))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet.append([])
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0B4778")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[4].height = 42

    for row_number, person in enumerate(personnel, start=5):
        personnel_id = person.get("id")
        can_enter = bool(can_enter_map.get(personnel_id, True))
        restriction_reasons = []
        if person.get("entry_blocked") is True or person.get("is_blocked") is True or person.get("blocked") is True:
            manual_reason = person.get("entry_block_reason") or person.get("block_reason")
            restriction_reasons.append(
                f"Yönetici tarafından engellenmiş{f': {manual_reason}' if manual_reason else ''}"
            )

        assignment_start = parse_dt_safe(person.get("assignment_start"))
        assignment_end = parse_dt_safe(person.get("assignment_end"))
        if assignment_start and assignment_start > now:
            restriction_reasons.append("Görevlendirme henüz başlamamış")
        if assignment_end and assignment_end < now:
            restriction_reasons.append("Görevlendirme tarihi sona ermiş")

        for document_type in mandatory_document_types:
            document = document_by_person.get((personnel_id, document_type.get("id")))
            document_name = document_type.get("name_tr") or document_type.get("name_en") or "Zorunlu evrak"
            if not document:
                restriction_reasons.append(f"Zorunlu belge eksik: {document_name}")
                continue
            expiry_date = parse_dt_safe(document.get("expiry_date"))
            if not expiry_date or expiry_date < now:
                restriction_reasons.append(f"Belge süresi dolmuş: {document_name}")

        row = [
            _excel_text(person.get("full_name")),
            _excel_text(person.get("tc_number")),
            _excel_text(person.get("company")),
            _excel_text(person.get("phone")),
            _excel_text(person.get("license_plate")),
            _excel_date(person.get("assignment_start")),
            _excel_date(person.get("assignment_end")),
            "Uygun" if can_enter else "Kısıtlı",
            "; ".join(restriction_reasons),
        ]
        sheet.append(row)

        sheet.cell(row_number, 6).number_format = "dd.mm.yyyy"
        sheet.cell(row_number, 7).number_format = "dd.mm.yyyy"

        column = len(base_headers) + 1
        for document_type in document_types:
            document = document_by_person.get((personnel_id, document_type.get("id")))
            expiry_value = document.get("expiry_date") if document else None
            expiry_cell = sheet.cell(row_number, column)
            days_cell = sheet.cell(row_number, column + 1)
            status_cell = sheet.cell(row_number, column + 2)

            if expiry_value:
                expiry_cell.value = _excel_date(expiry_value)
                if isinstance(expiry_cell.value, datetime):
                    expiry_cell.number_format = "dd.mm.yyyy"

            expiry_ref = expiry_cell.coordinate
            days_ref = days_cell.coordinate
            warning_days = int(document_type.get("warning_days") or 30)
            days_cell.value = f'=IF({expiry_ref}="","",{expiry_ref}-TODAY())'
            status_cell.value = (
                f'=IF({expiry_ref}="","Eksik",'
                f'IF({days_ref}<0,"Süresi Doldu",'
                f'IF({days_ref}<={warning_days},"Yaklaşıyor","Geçerli")))'
            )
            days_cell.number_format = "0"
            column += 3

    max_row = max(sheet.max_row, 5)
    sheet.auto_filter.ref = f"A4:{last_column}{max_row}"

    widths = [28, 16, 24, 17, 14, 20, 20, 15, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index in range(len(base_headers) + 1, len(headers) + 1, 3):
        sheet.column_dimensions[get_column_letter(index)].width = 24
        sheet.column_dimensions[get_column_letter(index + 1)].width = 13
        sheet.column_dimensions[get_column_letter(index + 2)].width = 16

    for row in sheet.iter_rows(min_row=5, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E2E8F0"))

    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    sheet.conditional_formatting.add(
        f"H5:H{max_row}",
        FormulaRule(formula=["$H5=\"Uygun\""], fill=green_fill),
    )
    sheet.conditional_formatting.add(
        f"H5:H{max_row}",
        FormulaRule(formula=["$H5=\"Kısıtlı\""], fill=red_fill),
    )
    for status_column in range(len(base_headers) + 3, len(headers) + 1, 3):
        letter = get_column_letter(status_column)
        target = f"{letter}5:{letter}{max_row}"
        sheet.conditional_formatting.add(target, FormulaRule(formula=[f'{letter}5="Geçerli"'], fill=green_fill))
        sheet.conditional_formatting.add(target, FormulaRule(formula=[f'{letter}5="Yaklaşıyor"'], fill=yellow_fill))
        sheet.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'OR({letter}5="Süresi Doldu",{letter}5="Eksik")'], fill=red_fill),
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"Clear2Work_Personel_Yedegi_{export_time.strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{personnel_id}")
async def get_personnel_detail(personnel_id: str, current_user: dict = Depends(get_current_user)):
    if DEMO_MODE:
        demo_map = {
            "demo_p_0": {"id": "demo_p_0", "full_name": "Ahmet Yılmaz", "tc_number": "12345678900", "company": "Demolife İnşaat", "phone": "+90 555 111 22 33", "assignment_end": (datetime.now() + timedelta(days=60)).isoformat()},
            "demo_p_1": {"id": "demo_p_1", "full_name": "Mehmet Demir", "tc_number": "12345678901", "company": "Demolife İnşaat", "phone": "+90 555 222 33 44", "assignment_end": (datetime.now() + timedelta(days=45)).isoformat()},
            "demo_p_2": {"id": "demo_p_2", "full_name": "Ayşe Kaya", "tc_number": "12345678902", "company": "Yıldız Liman", "phone": "+90 555 333 44 55", "assignment_end": (datetime.now() + timedelta(days=10)).isoformat()},
            "demo_p_3": {"id": "demo_p_3", "full_name": "Fatma Çelik", "tc_number": "12345678903", "company": "Yıldız Liman", "phone": "+90 555 444 55 66", "assignment_end": (datetime.now() - timedelta(days=5)).isoformat()},
            "demo_p_4": {"id": "demo_p_4", "full_name": "Can Özkan", "tc_number": "12345678904", "company": "Atlas Tersane", "phone": "+90 555 555 66 77", "assignment_end": (datetime.now() + timedelta(days=90)).isoformat()},
        }
        personnel = demo_map.get(personnel_id)
        if not personnel:
            raise HTTPException(status_code=404, detail="Personnel not found")

        now = datetime.now(timezone.utc)
        assignment_expired = False
        if personnel.get("assignment_end"):
            ae = datetime.fromisoformat(personnel["assignment_end"])
            if ae.tzinfo is None:
                ae = ae.replace(tzinfo=timezone.utc)
            if ae < now:
                assignment_expired = True

        demo_docs = [
            {
                "id": f"doc_{personnel_id}_1",
                "personnel_id": personnel_id,
                "document_type_id": "dt_1",
                "expiry_date": (now + timedelta(days=120)).isoformat(),
                "notes": "Demo belge",
                "document_type": {"id": "dt_1", "name_tr": "İş Güvenliği Eğitimi", "name_en": "Safety Training", "is_mandatory": True, "warning_days": 30},
                "status": "valid",
                "days_until_expiry": 120
            },
            {
                "id": f"doc_{personnel_id}_2",
                "personnel_id": personnel_id,
                "document_type_id": "dt_2",
                "expiry_date": (now + timedelta(days=200)).isoformat(),
                "notes": "Demo belge",
                "document_type": {"id": "dt_2", "name_tr": "Sağlık Raporu", "name_en": "Health Report", "is_mandatory": True, "warning_days": 30},
                "status": "valid",
                "days_until_expiry": 200
            },
        ]

        overall_status = "red" if assignment_expired else "green"
        status_reason = "assignment_expired" if assignment_expired else "all_valid"

        return {
            "personnel": personnel,
            "documents": demo_docs,
            "overall_status": overall_status,
            "status_reason": status_reason,
            "assignment_expired": assignment_expired,
            "assignment_not_started": False,
            "restriction_reasons": [{"code": "assignment_expired"}] if assignment_expired else [],
            "activity_summary": {
                "last_entry_at": None,
                "last_exit_at": None,
                "today_inside_seconds": 0,
                "last_gate": None,
            },
        }

    personnel = await db.personnel.find_one({"id": personnel_id}, {"_id": 0})
    if not personnel:
        raise HTTPException(status_code=404, detail="Personnel not found")

    documents = await db.personnel_documents.find({"personnel_id": personnel_id}, {"_id": 0}).to_list(100)

    doc_types = await db.document_types.find({}, {"_id": 0}).to_list(100)
    doc_types_map = {dt["id"]: dt for dt in doc_types}

    now = datetime.now(timezone.utc)
    enriched_docs = []
    for doc in documents:
        doc_type = doc_types_map.get(doc["document_type_id"])
        if doc_type:
            expiry = datetime.fromisoformat(doc["expiry_date"]) if isinstance(doc["expiry_date"], str) else doc["expiry_date"]
            if expiry.tzinfo is None:
                expiry = expiry.replace(tzinfo=timezone.utc)
            days_until_expiry = (expiry - now).days

            if days_until_expiry < 0:
                status_ = "expired"
            elif days_until_expiry <= doc_type["warning_days"]:
                status_ = "warning"
            else:
                status_ = "valid"

            enriched_docs.append({
                **doc,
                "document_type": doc_type,
                "status": status_,
                "days_until_expiry": days_until_expiry
            })

    restriction_reasons = []
    assignment_expired = False
    assignment_not_started = False
    if personnel.get("assignment_start"):
        assignment_start = datetime.fromisoformat(personnel["assignment_start"]) if isinstance(personnel["assignment_start"], str) else personnel["assignment_start"]
        if assignment_start.tzinfo is None:
            assignment_start = assignment_start.replace(tzinfo=timezone.utc)
        assignment_not_started = assignment_start > now
        if assignment_not_started:
            restriction_reasons.append({"code": "assignment_not_started"})

    if personnel.get("assignment_end"):
        assignment_end = datetime.fromisoformat(personnel["assignment_end"]) if isinstance(personnel["assignment_end"], str) else personnel["assignment_end"]
        if assignment_end.tzinfo is None:
            assignment_end = assignment_end.replace(tzinfo=timezone.utc)
        if assignment_end < now:
            assignment_expired = True
            restriction_reasons.append({"code": "assignment_expired"})

    if personnel.get("entry_blocked") is True or personnel.get("is_blocked") is True or personnel.get("blocked") is True:
        restriction_reasons.append({
            "code": "admin_blocked",
            "detail": personnel.get("block_reason") or personnel.get("entry_block_reason") or ""
        })

    uploaded_type_ids = {d.get("document_type_id") for d in enriched_docs}
    missing_mandatory = [
        dt for dt in doc_types
        if dt.get("is_mandatory") and dt.get("id") not in uploaded_type_ids
    ]
    if missing_mandatory:
        restriction_reasons.append({
            "code": "mandatory_document_missing",
            "documents": [dt.get("name_tr") or dt.get("name_en") or "" for dt in missing_mandatory]
        })

    expired_mandatory = [
        d for d in enriched_docs
        if d.get("document_type", {}).get("is_mandatory") and d.get("status") == "expired"
    ]
    if expired_mandatory:
        restriction_reasons.append({
            "code": "document_expired",
            "documents": [
                d.get("document_type", {}).get("name_tr")
                or d.get("document_type", {}).get("name_en")
                or ""
                for d in expired_mandatory
            ]
        })

    if restriction_reasons:
        overall_status = "red"
        status_reason = restriction_reasons[0]["code"]
    else:
        mandatory_docs = [d for d in enriched_docs if d["document_type"]["is_mandatory"]]
        has_warning = any(d["status"] == "warning" for d in mandatory_docs)

        if has_warning:
            overall_status = "yellow"
            status_reason = "warning_documents"
        else:
            overall_status = "green"
            status_reason = "all_valid"

    activity_summary = await _activity_summary(personnel_id)

    return {
        "personnel": personnel,
        "documents": enriched_docs,
        "overall_status": overall_status,
        "status_reason": status_reason,
        "assignment_expired": assignment_expired,
        "assignment_not_started": assignment_not_started,
        "restriction_reasons": restriction_reasons,
        "activity_summary": activity_summary,
    }


@router.put("/{personnel_id}")
async def update_personnel(personnel_id: str, data: PersonnelCreate, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["admin"])

    update_data = data.dict(exclude_none=True)
    result = await db.personnel.update_one({"id": personnel_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Personnel not found")
    return {"message": "Personnel updated"}


@router.delete("/{personnel_id}")
async def delete_personnel(personnel_id: str, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["admin"])

    result = await db.personnel.delete_one({"id": personnel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Personnel not found")

    await db.personnel_documents.delete_many({"personnel_id": personnel_id})
    return {"message": "Personnel deleted"}


@router.post("/bulk-import")
async def bulk_import_personnel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["admin"])

    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="File must be Excel or CSV format")

    try:
        contents = await file.read()
        if file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents), sep=None, engine="python")
        else:
            df = pd.read_excel(io.BytesIO(contents))

        df = df.dropna(how="all")

        # Clean BOM characters (ï»¿ or \ufeff) and strip whitespace from columns
        df.columns = [str(c).replace('\ufeff', '').replace('ï»¿', '').strip() for c in df.columns]

        # Map English template columns to Turkish if the template was downloaded in English
        column_mapping = {
            "Full Name": "Ad Soyad",
            "Company": "Şirket",
            "Assignment Start": "Görev Başlangıç",
            "Assignment End": "Görev Bitiş",
            "ID Number": "TC Kimlik No",
            "Phone": "Telefon",
            "License Plate": "Plaka",
            "Photo URL": "Fotoğraf URL"
        }
        df.rename(columns=column_mapping, inplace=True)

        doc_types = await db.document_types.find({}, {"_id": 0}).to_list(100)
        doc_type_map_tr = {dt["name_tr"]: dt["id"] for dt in doc_types}
        doc_type_map_en = {dt["name_en"]: dt["id"] for dt in doc_types}

        required_columns = ["Ad Soyad", "Şirket"]
        for col in required_columns:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        imported_count = 0
        skipped_count = 0
        errors = []

        def convert_date(date_str):
            if not date_str or date_str == "nan":
                return None
            if "00.01.1900" in date_str or "01.01.1900" in date_str:
                return None
            try:
                if "." in date_str:
                    parts = date_str.split(".")
                    if len(parts) == 3:
                        day, month, year = parts
                        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                return date_str
            except Exception:
                return None

        personnel_batch = []
        document_batch = []
        BATCH_SIZE = 100

        for index, row in df.iterrows():
            try:
                full_name = str(row.get("Ad Soyad", "")).strip()
                company = str(row.get("Şirket", "")).strip()

                if not full_name or not company or full_name == "nan" or company == "nan":
                    skipped_count += 1
                    continue

                assignment_start = str(row.get("Görev Başlangıç", "")).strip() if pd.notna(row.get("Görev Başlangıç")) else None
                assignment_end = str(row.get("Görev Bitiş", "")).strip() if pd.notna(row.get("Görev Bitiş")) else None
                assignment_start = convert_date(assignment_start) if assignment_start else None
                assignment_end = convert_date(assignment_end) if assignment_end else None

                tc_number = str(row.get("TC Kimlik No", "")).strip()
                if not tc_number:
                    tc_number = f"AUTO_{int(datetime.now(timezone.utc).timestamp() * 1000)}_{index}"

                personnel_id = f"{new_id('personnel')}_{index}"
                personnel_doc = {
                    "id": personnel_id,
                    "full_name": full_name,
                    "tc_number": tc_number,
                    "company": company,
                    "phone": str(row.get("Telefon", "")).strip() if pd.notna(row.get("Telefon")) else None,
                    "license_plate": str(row.get("Plaka", "")).strip() if pd.notna(row.get("Plaka")) else None,
                    "photo_url": str(row.get("Fotoğraf URL", "")).strip() if pd.notna(row.get("Fotoğraf URL")) else None,
                    "assignment_start": assignment_start,
                    "assignment_end": assignment_end,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }

                existing = await db.personnel.find_one({"full_name": full_name, "company": company})
                if existing:
                    errors.append(f"Row {index + 2}: {full_name} already exists")
                    skipped_count += 1
                    continue

                personnel_batch.append(personnel_doc)

                for col in df.columns:
                    if col not in required_columns and col not in ["Ad Soyad", "TC Kimlik No", "Telefon", "Plaka", "Fotoğraf URL", "Görev Başlangıç", "Görev Bitiş"]:
                        doc_type_id = doc_type_map_tr.get(col) or doc_type_map_en.get(col)
                        if doc_type_id and pd.notna(row.get(col)):
                            expiry_date_raw = str(row[col]).strip()
                            if expiry_date_raw and expiry_date_raw != "" and expiry_date_raw.lower() != "nan":
                                expiry_date = convert_date(expiry_date_raw)
                                if expiry_date:
                                    document_batch.append({
                                        "id": f"{new_id('doc')}_{index}_{col}",
                                        "personnel_id": personnel_id,
                                        "document_type_id": doc_type_id,
                                        "expiry_date": expiry_date,
                                        "notes": "Excel'den aktarıldı",
                                        "created_at": datetime.now(timezone.utc).isoformat()
                                    })

                imported_count += 1

                # Batch insert (her 100 kayıtta bir yaz)
                if len(personnel_batch) >= BATCH_SIZE:
                    await db.personnel.insert_many(personnel_batch)
                    personnel_batch = []
                if len(document_batch) >= BATCH_SIZE:
                    await db.personnel_documents.insert_many(document_batch)
                    document_batch = []

            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue

        # Kalan batch'leri yaz
        if personnel_batch:
            await db.personnel.insert_many(personnel_batch)
        if document_batch:
            await db.personnel_documents.insert_many(document_batch)

        return {"message": "Import completed", "imported": imported_count, "skipped": skipped_count, "errors": errors}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


@router.post("/bulk-delete")
async def bulk_delete_personnel(payload: BulkDeleteRequest, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["admin"])

    if not payload.ids:
        return {"deleted_count": 0}

    personnel_result = await db.personnel.delete_many({"id": {"$in": payload.ids}})
    await db.personnel_documents.delete_many({"personnel_id": {"$in": payload.ids}})
    return {"deleted_count": personnel_result.deleted_count}
